# -*- coding: utf-8 -*-
"""
build_data.py — Convierte data/Base - Prevalencia y dependencia.xlsx en web/data.js

web/data.js es GENERADO: no editarlo a mano. El dashboard lo consume y todos
los filtros se recalculan en el navegador a partir de los microdatos.

Uso:
    python scripts/validate.py && python scripts/build_data.py
"""
import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "data" / "Base - Prevalencia y dependencia.xlsx"
OUT = ROOT / "web" / "data.js"

# --- Entrevistas Confianza (pestaña "Entrevista de actores clave") ----------
# Base REAL: dashboard cualitativo (hoja 'Dashboard - Entrevistas', que ya trae
# las columnas calculadas: Respuesta del actor, Modalidad, Dummy realizada, etc.)
EXCEL_CUALI = ROOT / "data" / "APOYO Consultoría - Dashboard - seguimiento piloto cuali.xlsx"
OUT_CONF = ROOT / "web" / "data-confianza.js"
# (Se conserva por compatibilidad con build_gf, que aún usa la base de ejemplo.)
EXCEL_CONF = ROOT / "data" / "Base - Confianza y desempeño.xlsx"

# Normaliza el estado 'Respuesta del actor' (corrige mayúsculas/acentos)
EST_NORM = {
    "sin respuesta": "Sin respuesta", "aceptada": "Aceptada",
    "en coordinación": "En coordinación", "en coordinacion": "En coordinación",
    "rechazada": "Rechazada", "realizada completa": "Realizada completa",
}

# --- Grupos focales (pestaña "Grupos focales: Dependencia y Desempeño") -----
# Se alimenta de la hoja 'GF' del mismo Excel de Confianza/Desempeño.
OUT_GF = ROOT / "web" / "data-gf.js"

# --- Registro diario de encuestadoras (subpestaña de Prevalencia) -----------
# Fuente: extracto liviano de las DOS bases crudas (lo genera
# scripts/extraer_registro.py). Reemplaza al antiguo
# "Registro diario de encuestadoras.xlsx".
REG_EXTRACT = ROOT / "data" / "Registro - bases crudas.csv"
OUT_REG = ROOT / "web" / "data-registro.js"
# Fuente oficial de nombres/DDJJ por código de persona (evita el roster manual
# del Excel, que puede tener códigos mal escritos u omisiones).
EXCEL_BDPERSONAL = ROOT / "data" / "BD Personal.xlsx"

# --- Expedientes: Desempeño (pestaña "Expedientes: Desempeño") ---------------
# Base REAL: hoja 'Dashboard' del Excel de expedientes, que ya trae una fila por
# expediente revisado con las columnas calculadas (Revisor, Distrito Judicial,
# Tipo de caso, Minutos de revisión y fechas). Reproduce el Power BI.
EXCEL_EXP = ROOT / "data" / "Base - Expedientes.xlsx"
OUT_EXP = ROOT / "web" / "data-expedientes.js"

# --- Entrevistas a sobrevivientes (pestaña "Entrevistas a sobrevivientes") ---
# Base: hoja 'Dashboard - Sobrevivientes' del mismo Excel cuali (una fila por
# sobreviviente identificada/contactada; trae Distrito Judicial, Respuesta del
# actor, Resultado, Modalidad). La meta es FIJA (Tabla 9) y se define en el HTML.
OUT_SOB = ROOT / "web" / "data-sobrevivientes.js"


def _ddjj_norm(distrito):
    """Normaliza DISTRITO de BD Personal a la grafía de DDJJ del dashboard."""
    if not distrito:
        return ""
    key = " ".join(str(distrito).strip().upper().split())
    mapa = {
        "UCAYALI": "Ucayali", "LIMA NORTE": "Lima Norte",
        "APURÍMAC": "Apurímac", "APURIMAC": "Apurímac",
        "CAJAMARCA": "Cajamarca", "LORETO": "Loreto", "AYACUCHO": "Ayacucho",
        "HUÁNUCO": "Huánuco", "HUANUCO": "Huánuco",
        "PUENTE PIEDRA - VENTANILLA": "Puente Piedra-Ventanilla",
        "PUENTE PIEDRA-VENTANILLA": "Puente Piedra-Ventanilla",
    }
    return mapa.get(key, key.title())


def _leer_bdpersonal():
    """code -> {nombre, ddjj} desde BD Personal.xlsx (hoja 'Exported')."""
    mapa = {}
    if not EXCEL_BDPERSONAL.exists():
        return mapa
    wb = openpyxl.load_workbook(EXCEL_BDPERSONAL, data_only=True, read_only=True)
    ws = wb["Exported"] if "Exported" in wb.sheetnames else wb[wb.sheetnames[0]]
    filas = ws.iter_rows(min_row=2, values_only=True)
    for r in filas:
        if not r or r[0] is None:
            continue
        cod = _norm_cod(r[0])
        if not cod:
            continue
        nombre = str(r[1]).strip() if len(r) > 1 and r[1] else cod
        ddjj = _ddjj_norm(r[3]) if len(r) > 3 else ""
        mapa[cod] = {"nombre": nombre, "ddjj": ddjj}
    wb.close()
    return mapa

# p800resultado (código) -> etiqueta (según hoja Apoyo del Registro)
RES_MAP = {
    "01": "Completa", "02": "Cita - Aplazada", "03": "Ausente", "04": "Discapacitada",
    "05": "No habita población objetivo en la vivienda", "06": "Rechazada", "07": "Incompleta",
}
# Columnas del extracto liviano (data/Registro - bases crudas.csv). El grupo
# poblacional (Adolescentes/Mujeres) ya viene como una columna del extracto.
REG_COLNAMES = {
    "grupo": "grupo",
    "cod": "codencu",
    "enc": "encuestador",   # nombre de la encuestadora registrado en la base
    "fi": "fechainicioenc",
    "ff": "fechafinenc",
    "res": "p800resultado",
}


def _norm_cod(x):
    """Normaliza un código de persona a 2 dígitos extrayendo sus dígitos.

    Maneja casos como 8 -> '08', '02' -> '02' y 'E44' -> '44' (código con
    prefijo de rol mal digitado en la base).
    """
    if x is None:
        return ""
    s = str(x).strip()
    if s in ("", "."):
        return ""
    digitos = "".join(ch for ch in s if ch.isdigit())
    if not digitos:
        return ""
    return f"{int(digitos):02d}"


def _norm_name_key(s):
    """Clave canónica de un nombre para agrupar a la misma persona.

    Quita acentos y caracteres corruptos (la columna 'encuestador' de la base
    trae mojibake: la Ñ aparece como el carácter de reemplazo U+FFFD), pasa a
    MAYÚSCULAS y colapsa espacios. Así un mismo nombre registrado bajo dos
    códigos distintos (p. ej. ROSA ELVIRA con 40 y 64) se agrupa en una sola
    persona.
    """
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = "".join(ch if (ch.isalnum() or ch.isspace()) else " " for ch in s)
    return " ".join(s.upper().split())

# Rol crudo -> etiqueta corta usada como "Tipo de actor" en el dashboard
ROL_MAP = {
    "Acompañante de víctimas": "Acompañantes",
    "Actor local o comunitario": "Actores locales",
    "Funcionario del Sistema de Justicia": "Funcionarios del Sistema",
    "Multilateral": "Multilaterales",
    "Prensa": "Prensa",
}
MODALIDAD_DEFAULT = "A distancia asistido"


def hora_a_min(s):
    """'11:05:24' -> minutos desde medianoche (float) o None."""
    try:
        h, m, sec = str(s).split(":")
        return int(h) * 60 + int(m) + int(sec) / 60.0
    except (ValueError, AttributeError):
        return None


def _clean(x):
    """Normaliza celdas: '.', None y vacíos -> '' (cadena vacía)."""
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s in (".", "None", "nan") else s


def estado_coordinacion(contacto, coord, resultado):
    """Regla confirmada para 'Respuesta del actor' (estado de coordinación).

    Precedencia:
      1. Realizada completa : Resultado == 'Completa'
      2. Rechazada          : Coordinación == 'Rechazada'
      3. En coordinación     : Coordinación == 'Coordinada'
      4. Aceptada            : Contacto == 'Contactado' (contactado, aún sin coordinar)
      5. Sin respuesta       : resto (No contactado)
    """
    if resultado == "Completa":
        return "Realizada completa"
    if coord == "Rechazada":
        return "Rechazada"
    if coord == "Coordinada":
        return "En coordinación"
    if contacto == "Contactado":
        return "Aceptada"
    return "Sin respuesta"


def _ambito_distrito(dj):
    """Del campo 'Distrito judicial' deriva (distrito_limpio, ámbito).

    Valores típicos: 'Nacional' o 'Distrito Judicial: Huánuco'.
    """
    s = _clean(dj)
    if not s:
        return None, None
    if s.lower() == "nacional":
        return "Nacional", "Nacional"
    if ":" in s:
        return s.split(":", 1)[1].strip(), "Local"
    return s, "Local"


def build_confianza() -> None:
    """Genera web/data-confianza.js desde la base real (dashboard cualitativo).

    Lee la hoja 'Dashboard - Entrevistas' filtrando 'Entrevista de actores clave'.
    Esa hoja ya trae las columnas calculadas del Power BI (Respuesta del actor,
    Modalidad, Dummy realizada, Duración), así que se leen directamente y el
    resultado cuadra con el Power BI. El universo son las entrevistas con un
    estado de coordinación definido (las que aún no se contactan quedan fuera).
    """
    if not EXCEL_CUALI.exists():
        print(f"AVISO: no se encontró {EXCEL_CUALI.name}; se omite data-confianza.js")
        return

    ent = pd.read_excel(EXCEL_CUALI, sheet_name="Dashboard - Entrevistas")
    ent.columns = [str(c).strip() for c in ent.columns]

    # Universo: entrevistas a actores clave
    ent = ent[ent["Tipo de entrevistas"].map(lambda x: _clean(x).lower()) == "entrevista de actores clave"].copy()

    registros = []
    for _, r in ent.iterrows():
        estado_raw = _clean(r.get("Respuesta del actor"))
        if not estado_raw:
            continue  # sin estado de coordinación aún -> fuera de la dona
        estado = EST_NORM.get(estado_raw.lower(), estado_raw)
        distrito, ambito = _ambito_distrito(r.get("Distrito judicial"))
        dur = r.get("Duración total (min)")
        try:
            dur = float(str(dur).strip())
            if pd.isna(dur):          # celdas vacías -> NaN -> None
                dur = None
        except (ValueError, TypeError, AttributeError):
            dur = None
        registros.append([
            distrito,                                               # 0 distrito (limpio)
            _clean(r.get("Tipo de actor")) or None,                 # 1 tipo_actor
            ambito,                                                 # 2 ámbito (Local/Nacional)
            _clean(r.get("Modalidad")) or None,                     # 3 modalidad
            estado,                                                 # 4 estado (Respuesta del actor)
            1 if estado == "Realizada completa" else 0,             # 5 realizada
            dur,                                                    # 6 duración (min)
        ])

    data = {
        "meta": {
            "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fuente": EXCEL_CUALI.name,
            "n_registros": len(registros),
            "modalidad_constante": False,
            "campos": ["distrito", "tipo_actor", "ambito", "modalidad", "estado", "realizada", "dur_min"],
        },
        "registros": registros,
    }
    js = "// GENERADO por scripts/build_data.py — NO EDITAR A MANO\n"
    js += "const DATA_CONF = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUT_CONF.write_text(js, encoding="utf-8")

    real = [x for x in registros if x[5] == 1]
    durs = [x[6] for x in real if x[6] is not None]
    print(f"OK -> {OUT_CONF.relative_to(ROOT)}")
    print(f"   actores clave (con estado): {len(registros)} | realizadas: {len(real)} | "
          f"duración prom: {round(sum(durs)/len(durs),1) if durs else '—'}")


def _fecha_iso(v):
    """Devuelve 'YYYY-MM-DD' desde un valor de fecha de pandas/Excel, o None."""
    try:
        ts = pd.to_datetime(v, errors="coerce", dayfirst=True)
        return ts.strftime("%Y-%m-%d") if pd.notna(ts) else None
    except Exception:  # noqa: BLE001
        return None


def build_gf() -> None:
    """Genera web/data-gf.js desde la hoja 'GF' (Grupos focales).

    Replica la tabla 'Dashboard - Reunión (2)' del Power BI. Es tolerante con
    columnas que la base de ejemplo no trae todavía ('N° de Grupo Focal',
    'Fecha de grupo focal', 'Duración'): usa fallbacks razonables.
    """
    if not EXCEL_CONF.exists():
        print(f"AVISO: no se encontró {EXCEL_CONF.name}; se omite data-gf.js")
        return

    gf = pd.read_excel(EXCEL_CONF, sheet_name="GF")
    gf.columns = [str(c).strip() for c in gf.columns]
    cols = set(gf.columns)

    tiene_ngf = "N° de Grupo Focal" in cols
    tiene_fecha_gf = "Fecha de grupo focal" in cols
    tiene_dur = "Duración" in cols
    # Solo filas con algún contenido
    gf = gf.dropna(how="all")

    registros = []
    for _, r in gf.iterrows():
        contacto = _clean(r.get("Contacto"))
        coord = _clean(r.get("Coordinación"))
        asis = _clean(r.get("Asistencia"))
        ta = _clean(r.get("Tipo de actor"))
        ngf = _clean(r.get("N° de Grupo Focal")) if tiene_ngf else ""
        fecha_src = r.get("Fecha de grupo focal") if tiene_fecha_gf else r.get("Fecha de contacto")
        dur = r.get("Duración") if tiene_dur else None
        registros.append([
            ngf or "Grupo focal 1",                                  # 0 ngf
            _clean(r.get("Distrito Judicial")) or None,              # 1 distrito
            ROL_MAP.get(ta, ta) or None,                             # 2 tipo_actor
            estado_coordinacion(contacto, coord, ""),                # 3 estado
            1 if contacto == "Contactado" else 0,                    # 4 contactado
            1 if coord == "Coordinada" else 0,                       # 5 confirmado
            1 if asis == "Asistió" else 0,                           # 6 asistio
            asis or None,                                            # 7 asistencia (leyenda)
            _fecha_iso(fecha_src),                                   # 8 fecha
            float(dur) if isinstance(dur, (int, float)) and pd.notna(dur) else None,  # 9 dur
        ])

    data = {
        "meta": {
            "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fuente": EXCEL_CONF.name + " (hoja GF)",
            "n_registros": len(registros),
            "tiene_ngf": tiene_ngf,
            "tiene_fecha_gf": tiene_fecha_gf,
            "tiene_duracion": tiene_dur,
            "campos": ["ngf", "distrito", "tipo_actor", "estado", "contactado",
                       "confirmado", "asistio", "asistencia", "fecha", "dur_min"],
        },
        "registros": registros,
    }
    js = "// GENERADO por scripts/build_data.py — NO EDITAR A MANO\n"
    js += "const DATA_GF = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUT_GF.write_text(js, encoding="utf-8")
    print(f"OK -> {OUT_GF.relative_to(ROOT)}")
    print(f"   contactos: {len(registros)} | contactados: {sum(x[4] for x in registros)} | "
          f"asistentes: {sum(x[6] for x in registros)}")


def build_registro() -> None:
    """Genera web/data-registro.js desde el extracto de las bases crudas.

    Microdatos por encuesta (código de encuestadora, grupo, resultado, fecha de
    inicio y de fin) leídos del extracto 'Registro - bases crudas.csv', que
    scripts/extraer_registro.py arma a partir de las dos bases crudas de campo
    (DATA_ADOLESCENTES y DATA_MUJERES). Los nombres y DDJJ se toman de
    BD Personal.xlsx (fuente oficial) por código, de modo que el total cuadra con
    el avance general. El navegador arma la matriz encuestadora x fecha y aplica
    los filtros (Grupo, Resultado, Distrito Judicial, base de fecha Inicio/Fin).
    """
    if not REG_EXTRACT.exists():
        print(f"AVISO: no se encontró {REG_EXTRACT.name}; se omite data-registro.js")
        print("   (genera el extracto con: python scripts/extraer_registro.py)")
        return

    import csv as _csv
    from collections import Counter, defaultdict

    # --- BD Personal: solo se usa para el DDJJ (por código) y la grafía limpia
    #     del nombre. La IDENTIDAD de quién hizo cada encuesta sale de la base. ---
    bd = _leer_bdpersonal()                       # cod -> {nombre, ddjj}
    bd_ddjj_por_nombre = {}                        # nombre_norm -> ddjj (respaldo)
    for _c, _i in bd.items():
        _k = _norm_name_key(_i.get("nombre", ""))
        if _k and _i.get("ddjj") and _k not in bd_ddjj_por_nombre:
            bd_ddjj_por_nombre[_k] = _i["ddjj"]

    # --- Microdatos del extracto (una fila por encuesta) ---
    # La persona se identifica por el NOMBRE 'encuestador' de la base (no por el
    # código): así un mismo nombre con dos códigos se agrupa, y un código que dos
    # personas usaron por error no las mezcla.
    registros = []                                 # [persona, grupo, res, fi, ff]
    codes_por_persona = defaultdict(Counter)       # persona -> Counter(cod)
    nombres_por_persona = defaultdict(Counter)     # persona -> Counter(nombre crudo)
    ddjj_por_persona = defaultdict(Counter)        # persona -> Counter(ddjj por código)
    with open(REG_EXTRACT, encoding="utf-8-sig", newline="") as fh:
        rd = _csv.DictReader(fh)
        # Normaliza encabezados a las claves esperadas (tolerante a mayúsculas).
        campo = {k: None for k in REG_COLNAMES}
        for col in (rd.fieldnames or []):
            low = str(col).strip().lower()
            for k, name in REG_COLNAMES.items():
                if low == name:
                    campo[k] = col
        for row in rd:
            cod = _norm_cod(row.get(campo["cod"])) if campo["cod"] else ""
            nombre_base = _clean(row.get(campo["enc"])) if campo["enc"] else ""
            # Identidad de la persona: nombre de la base; si falta, el código.
            persona = _norm_name_key(nombre_base) or (f"COD{cod}" if cod else "")
            if not persona:
                continue
            grupo = _clean(row.get(campo["grupo"])) if campo["grupo"] else ""
            rv = row.get(campo["res"]) if campo["res"] else None
            res = RES_MAP.get(str(rv).strip().zfill(2) if rv not in (None, "") else "", None)
            fi = _fecha_iso(row.get(campo["fi"])) if campo["fi"] else None
            ff = _fecha_iso(row.get(campo["ff"])) if campo["ff"] else None
            if res is None and fi is None and ff is None:
                continue
            registros.append([persona, grupo, res, fi, ff])
            if cod:
                codes_por_persona[persona][cod] += 1
                d = bd.get(cod, {}).get("ddjj", "")
                if d:
                    ddjj_por_persona[persona][d] += 1
            if nombre_base:
                nombres_por_persona[persona][nombre_base] += 1

    # --- Construir el mapa de encuestadoras (1 fila = 1 persona) ---
    #   cod    : código primario para mostrar y ordenar (el más frecuente)
    #   nombre : grafía limpia (BD Personal por código; si no, nombre de la base)
    #   ddjj   : DDJJ mayoritario según BD Personal (por código); respaldo por nombre
    encuestadoras = {}
    sin_ddjj = []
    for persona in {r[0] for r in registros}:
        cc = codes_por_persona[persona]
        modal_cod = min(cc, key=lambda c: (-cc[c], int(c))) if cc else ""

        dc = ddjj_por_persona[persona]
        if dc:
            ddjj = max(dc, key=lambda d: (dc[d], d))      # DDJJ mayoritario por código
        else:
            ddjj = bd_ddjj_por_nombre.get(persona, "")     # respaldo por nombre
        if not ddjj:
            ddjj = "(sin asignar)"
            sin_ddjj.append((persona, modal_cod))

        nombre = bd.get(modal_cod, {}).get("nombre", "") if modal_cod else ""
        if not nombre:
            nb = nombres_por_persona[persona].most_common(1)
            nombre = nb[0][0] if nb else persona
            nombre = " ".join(nombre.replace("\ufffd", "").split())  # limpia mojibake

        encuestadoras[persona] = {"cod": modal_cod or "—", "nombre": nombre, "ddjj": ddjj}

    if sin_ddjj:
        _lst = ["{} (cód. {})".format(n, c or "?") for n, c in sorted(sin_ddjj)]
        print("   AVISO: encuestadoras sin DDJJ en BD Personal: " + str(_lst))

    fechas = [r[3] for r in registros if r[3]] + [r[4] for r in registros if r[4]]
    data = {
        "meta": {
            "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fuente": "Bases crudas (Adolescentes y Mujeres)",
            "n_registros": len(registros),
            "fecha_min": min(fechas) if fechas else None,
            "fecha_max": max(fechas) if fechas else None,
            "campos": ["persona", "grupo", "resultado", "fecha_inicio", "fecha_fin"],
        },
        "encuestadoras": encuestadoras,
        "registros": registros,
    }
    js = "// GENERADO por scripts/build_data.py — NO EDITAR A MANO\n"
    js += "const DATA_REG = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUT_REG.write_text(js, encoding="utf-8")
    comp = sum(1 for r in registros if r[2] == "Completa")
    print(f"OK -> {OUT_REG.relative_to(ROOT)}")
    print(f"   encuestas: {len(registros)} | encuestadoras: {len(encuestadoras)} | completas: {comp}")


def _norm_revisor_map(nombres):
    """nombre_crudo -> nombre canónico, unificando variantes con/sin tilde.

    Ej.: 'Mayra Sanchez Hinojosa' y 'Mayra Sánchez Hinojosa' -> una sola opción.
    """
    import unicodedata
    from collections import Counter

    def key(s):
        return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower().strip()

    grupos = {}
    for n in nombres:
        grupos.setdefault(key(n), Counter())[n] += 1
    canon = {}
    for cnt in grupos.values():
        # preferir la variante con tildes (difiere de su forma ASCII) y más frecuente
        mejor = sorted(
            cnt.items(),
            key=lambda kv: (sum(1 for ch in kv[0] if ord(ch) > 127), kv[1]),
            reverse=True,
        )[0][0]
        for n in cnt:
            canon[n] = mejor
    return canon


def build_expedientes() -> None:
    """Genera web/data-expedientes.js desde la hoja 'Dashboard' del Excel de expedientes.

    Una fila por expediente revisado. El navegador aplica los filtros
    (Revisor / Tipo / Distrito Judicial) y recalcula: recuento de expedientes,
    expedientes revisados por día (Fecha-Inicio), tiempo promedio de revisión
    (Minutos) y las donas por Distrito Judicial y por Tipo de caso.
    """
    if not EXCEL_EXP.exists():
        print(f"AVISO: no se encontró {EXCEL_EXP.name}; se omite data-expedientes.js")
        return

    exp = pd.read_excel(EXCEL_EXP, sheet_name="Dashboard")
    exp.columns = [str(c).strip() for c in exp.columns]
    exp = exp.dropna(how="all")

    canon = _norm_revisor_map([_clean(v) for v in exp.get("Revisor", []) if _clean(v)])

    registros = []
    for _, r in exp.iterrows():
        cod = _clean(r.get("Código"))
        if not cod:
            continue
        rev_raw = _clean(r.get("Revisor"))
        revisor = canon.get(rev_raw, rev_raw) or None
        distrito = _clean(r.get("Distrito Judicial")) or None
        tipo = _clean(r.get("Tipo")) or None
        fecha = _fecha_iso(r.get("Fecha-Inicio"))
        try:
            mins = float(r.get("Minutos"))
            if pd.isna(mins):
                mins = None
        except (ValueError, TypeError):
            mins = None
        registros.append([cod, revisor, distrito, tipo, fecha, mins])

    data = {
        "meta": {
            "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fuente": EXCEL_EXP.name,
            "n_registros": len(registros),
            "campos": ["codigo", "revisor", "distrito", "tipo", "fecha", "minutos"],
        },
        "registros": registros,
    }
    js = "// GENERADO por scripts/build_data.py — NO EDITAR A MANO\n"
    js += "const DATA_EXP = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUT_EXP.write_text(js, encoding="utf-8")

    from collections import Counter
    mins = [r[5] for r in registros if r[5] is not None]
    prom = round(sum(mins) / len(mins)) if mins else "—"
    print(f"OK -> {OUT_EXP.relative_to(ROOT)}")
    print(f"   expedientes: {len(registros)} | revisores: {len(set(r[1] for r in registros))} | "
          f"tiempo prom: {prom} min")
    print(f"   distrito: {dict(Counter(r[2] for r in registros))} | "
          f"tipo: {dict(Counter(r[3] for r in registros))} | "
          f"por día: {dict(Counter(r[4] for r in registros))}")


def build_sobrevivientes() -> None:
    """Genera web/data-sobrevivientes.js desde la hoja 'Dashboard - Sobrevivientes'.

    Una fila por sobreviviente identificada/contactada. El navegador aplica los
    filtros (Ámbito / Distrito judicial) y recalcula: realizadas, duración
    promedio, % de avance vs meta (fija, Tabla 9), estado de coordinación,
    estado de las realizadas y entrevistas por distrito.
    """
    if not EXCEL_CUALI.exists():
        print(f"AVISO: no se encontró {EXCEL_CUALI.name}; se omite data-sobrevivientes.js")
        return
    try:
        sob = pd.read_excel(EXCEL_CUALI, sheet_name="Dashboard - Sobrevivientes")
    except Exception as e:  # noqa: BLE001  (la hoja puede no existir todavía)
        print(f"AVISO: no se pudo leer 'Dashboard - Sobrevivientes' ({type(e).__name__}); se omite.")
        return
    sob.columns = [str(c).strip() for c in sob.columns]

    registros = []
    for _, r in sob.iterrows():
        # Los encabezados de fila 1 son los originales; el contenido de las columnas
        # fue reorganizado. Mapeo observado:
        #   pandas "Contacto"               -> valor "Distrito Judicial: Callao"
        #   pandas "Fecha de primer contacto" -> estado: "En coordinación" / "Realizada completa"
        #   pandas "Respuesta del actor"    -> ¿Contactada? "Sí"/"No"
        #   pandas "Distrito Judicial"      -> nombre de orientadora
        distrito_raw = _clean(r.get("Contacto"))
        estado_raw   = _clean(r.get("Fecha de primer contacto"))
        contactada   = _clean(r.get("Respuesta del actor"))   # "Sí" / "No"
        orientadora  = _clean(r.get("Distrito Judicial"))     # nombre orientadora
        if not (distrito_raw or estado_raw or contactada):
            continue  # fila de relleno (vacía)
        # "Distrito Judicial: Huánuco" -> "Huánuco"; "Nacional" -> "Nacional"
        distrito = distrito_raw.split(":", 1)[1].strip() if ":" in distrito_raw else (distrito_raw or None)
        estado = EST_NORM.get(estado_raw.lower(), estado_raw) if estado_raw else None
        realizada = 1 if estado == "Realizada completa" else 0
        resultado = _clean(r.get("Resultado")) or None
        modalidad = _clean(r.get("Modalidad")) or None
        try:
            dur = float(r.get("Duración total (min)"))
            if pd.isna(dur):
                dur = None
        except (ValueError, TypeError):
            dur = None
        fecha = _fecha_iso(r.get("Fecha de inicio de entrevista"))
        registros.append([distrito, orientadora or None, estado, realizada, resultado, modalidad, dur, fecha])

    data = {
        "meta": {
            "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fuente": EXCEL_CUALI.name + " (hoja Dashboard - Sobrevivientes)",
            "n_registros": len(registros),
            "campos": ["distrito", "contacto", "estado", "realizada", "resultado", "modalidad", "duracion", "fecha"],
        },
        "registros": registros,
    }
    js = "// GENERADO por scripts/build_data.py — NO EDITAR A MANO\n"
    js += "const DATA_SOB = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUT_SOB.write_text(js, encoding="utf-8")
    from collections import Counter
    print(f"OK -> {OUT_SOB.relative_to(ROOT)}")
    print(f"   registros: {len(registros)} | realizadas: {sum(r[3] for r in registros)} | "
          f"distritos: {dict(Counter(r[0] for r in registros))}")


def main() -> None:
    base = pd.read_excel(EXCEL, sheet_name="Base (0)")
    cuotas = pd.read_excel(EXCEL, sheet_name="Cuotas")

    # --- Registros (una fila por encuesta) --------------------------------
    base = base[base["Resultado"].notna()].copy()
    base["fecha"] = pd.to_datetime(base["Fecha"], dayfirst=True, errors="coerce")

    ini = base["Inicio"].map(hora_a_min)
    fin = base["Fin"].map(hora_a_min)
    dur = fin - ini
    # Cruce de medianoche: duración negativa -> +24h
    dur = dur.where(dur.isna() | (dur >= 0), dur + 24 * 60)
    base["dur_min"] = dur.round(1)

    registros = []
    for _, r in base.iterrows():
        registros.append([
            r["fecha"].strftime("%Y-%m-%d") if pd.notna(r["fecha"]) else None,
            str(r["DDJJ"]) if pd.notna(r["DDJJ"]) else None,
            str(r["Grupo"]) if pd.notna(r["Grupo"]) else None,
            str(r["Sexo"]) if pd.notna(r["Sexo"]) and str(r["Sexo"]) != "." else None,
            int(r["Edad"]) if pd.notna(r["Edad"]) else None,
            str(r["Resultado"]),
            float(r["dur_min"]) if pd.notna(r["dur_min"]) else None,
        ])

    # --- Cuotas ------------------------------------------------------------
    cuotas_out = [
        {
            "ddjj": str(r["Distrito Judicial"]),
            "grupo": str(r["Grupo etario"]),
            "sexo": str(r["Sexo"]),
            "cuota": int(r["Cuota"]),
        }
        for _, r in cuotas.iterrows()
    ]

    data = {
        "meta": {
            "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fuente": EXCEL.name,
            "n_registros": len(registros),
            "cuota_total": int(cuotas["Cuota"].sum()),
            # Campos de cada registro, en orden:
            "campos": ["fecha", "ddjj", "grupo", "sexo", "edad", "resultado", "dur_min"],
        },
        "registros": registros,
        "cuotas": cuotas_out,
    }

    js = "// GENERADO por scripts/build_data.py — NO EDITAR A MANO\n"
    js += "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUT.write_text(js, encoding="utf-8")

    # Chequeo rápido en consola
    comp = base[base["Resultado"] == "Completa"]
    print(f"OK -> {OUT.relative_to(ROOT)}")
    print(f"   registros: {len(registros)} | completas: {len(comp)} | cuota total: {int(cuotas['Cuota'].sum())}")
    print(f"   avance: {len(comp) / cuotas['Cuota'].sum() * 100:.1f}%")

    # --- Pestaña Entrevistas Confianza (no detiene la build si falla) ---
    try:
        build_confianza()
    except Exception as e:  # noqa: BLE001
        print(f"AVISO: no se pudo generar data-confianza.js ({type(e).__name__}: {e})")

    # --- Pestaña Grupos focales (no detiene la build si falla) ---
    try:
        build_gf()
    except Exception as e:  # noqa: BLE001
        print(f"AVISO: no se pudo generar data-gf.js ({type(e).__name__}: {e})")

    # --- Subpestaña Registro diario de encuestadoras (no detiene la build) ---
    try:
        build_registro()
    except Exception as e:  # noqa: BLE001
        print(f"AVISO: no se pudo generar data-registro.js ({type(e).__name__}: {e})")

    # --- Pestaña Expedientes: Desempeño (no detiene la build) ---
    try:
        build_expedientes()
    except Exception as e:  # noqa: BLE001
        print(f"AVISO: no se pudo generar data-expedientes.js ({type(e).__name__}: {e})")

    # --- Pestaña Entrevistas a sobrevivientes (no detiene la build) ---
    try:
        build_sobrevivientes()
    except Exception as e:  # noqa: BLE001
        print(f"AVISO: no se pudo generar data-sobrevivientes.js ({type(e).__name__}: {e})")


if __name__ == "__main__":
    main()
